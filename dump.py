import sys
import json

from openpyxl import Workbook

'''
	Create list of dict headers
'''
def list_headers(json, headers):
	for key, value in json.items():
		headers.append(key)

		if isinstance(value, dict):
			list_headers(value, headers)
		elif isinstance(value, list):
			for val in value:
				if isinstance(val, dict):
					list_headers(val, headers)

'''
	Create list of values
'''
def dump_values(json):
	values = []

	for key, value in json.items():
		if isinstance(value, dict):
			values.append(None)
			values.extend(dump_values(value))
		elif isinstance(value, list):
			values.append(None)
			for val in value:
				if isinstance(val, dict):
					values.extend(dump_values(val))
				else:
					values.append(val)
		else:
			values.append(value)

	return values

'''
	Export as XLSX
'''
def export(ws, json):
	headers = []
	list_headers(list(json.values())[0], headers)

	for i, label in enumerate(headers):
		ws.cell(row = 2 + i, column = 1, value = label)

	for i, player in enumerate(json):
		ws.cell(row = 1, column = 2 + i, value = player)

		for j, val in enumerate(dump_values(json[player])):
			ws.cell(row = 2 + j, column = 2 + i, value = val)

'''
	Main
'''
def main(path):
	with open(path, 'r', encoding = 'utf8') as jfile:
		wb = Workbook()
		wb.active.title = 'Player Dump'

		export(wb['Player Dump'], json.load(jfile))

		wb.save(path.split('.')[0] + '.dump.xlsx')

if __name__ == '__main__':
	main(sys.argv[1])