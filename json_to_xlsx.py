import sys
import json
import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.borders import BORDER_THIN

'''
	Create list of dict headers
'''
def list_headers(json, headers, indent = 0):	
	for key, value in json.items():
		if 'data' in key:
			continue

		prv = headers[indent] and headers[indent][-1]
		if prv == None:
			headers[indent].pop(len(headers[indent]) - 1)
			if indent == 1:
				headers[2].pop(len(headers[2]) - 1)
			elif indent == 2:
				headers[0].pop(len(headers[0]) - 1)

		prv2 = headers[2] and headers[2][-1]
		if indent == 0 and prv2 is not None and headers[1]:
			headers[1].pop(len(headers[1]) - 1)

		for i in range(0, 3):
			headers[i].append(key if i == indent else None)


		if isinstance(value, dict):
			list_headers(value, headers, indent + 1)
		elif isinstance(value, list):
			for val in value:
				if isinstance(val, dict):
					list_headers(val, headers, indent + 1)

def dump_values(json):
	values = []

	for key, value in json.items():
		if 'data' in key:
			continue

		if isinstance(value, dict):
			values.append(None)
			values.extend(dump_values(value))
		elif isinstance(value, list):
			values.append(None)
			for val in value:
				if isinstance(val, dict):
					values.extend(dump_values(val))
				else:
					values.append(val)
		else:
			values.append(value)

	return values

'''
	Export as XLSX
'''
def export(ws, json):
	font_bold = Font(bold = True)
	font_italic = Font(italic = True)
	align_center = Alignment(horizontal = 'center')
	align_right = Alignment(horizontal = 'right')
	border_bottom = Border(bottom = Side(border_style = BORDER_THIN, color = 'FF000000'))

	ws.cell(row = 1, column = 1, value = 'name').font = font_bold

	headers = [[], [], []]
	list_headers(list(json.values())[0], headers)

	for i in range(0, len(headers[0])):
		c = ws.cell(row = 1, column = i + 2, value = headers[0][i])
		c.font = font_bold
		c.alignment = align_center
	for i in range(0, len(headers[1])):
		c = ws.cell(row = 2, column = i + 2, value = headers[1][i])
		c.alignment = align_center
	for i in range(0, len(headers[2])):
		c = ws.cell(row = 3, column = i + 2, value = headers[2][i])
		c.font = font_italic
		c.alignment = align_center

	for i in range(1, len(headers[0]) + 2):
		ws.cell(row = 3, column = i).border = border_bottom

	for i, player in enumerate(json):
		c = ws.cell(row = i + 4, column = 1, value = player)
		c.font = font_bold
		c.alignment = align_right

		col = 2
		for val in dump_values(json[player]):
			if val is not None:
				d = ws.cell(row = i + 4, column = col, value = val)
				d.alignment = align_center
				col += 1

'''
	Main
'''
def main(path):
	with open(path, 'r', encoding = 'utf8') as jfile:
		wb = None
		try:
			wb = openpyxl.load_workbook('template.xlsx')
		except:
			wb = Workbook()
			wb.active.title = 'Player Data'

		export(wb['Player Data'], json.load(jfile))

		wb.save(path.split('.')[0] + '.xlsx')

if __name__ == '__main__':
	main(sys.argv[1])